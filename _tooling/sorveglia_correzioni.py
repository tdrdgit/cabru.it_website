#!/usr/bin/env python3
"""Guarda ogni giorno se qualcuno ha scritto una correzione nel foglio dei testi.

Il foglio vive dentro la cartella del progetto, che sta su Google Drive: quando
qualcuno lo modifica online, la modifica torna sul Mac da sola. Quindi qui non
serve nessun accesso a Google — si legge il file come un file qualsiasi.

Avvisa su Telegram SOLO quando c'e' qualcosa di nuovo rispetto all'ultimo giro:
un avviso che arriva tutti i giorni uguale smette di essere letto.

    python3 sorveglia_correzioni.py           # il giro normale
    python3 sorveglia_correzioni.py --sempre  # avvisa anche se non e' cambiato niente
"""
import sys, os, json, hashlib, datetime

HERE = os.path.dirname(os.path.abspath(__file__))
_pylib = os.path.join(HERE, "pylib")
if os.path.isdir(_pylib): sys.path.insert(0, _pylib)
sys.path.insert(0, os.path.expanduser("~/.claude/scripts"))
import openpyxl

XLS   = os.path.join(HERE, "CABRU_testi_IT_da_correggere.xlsx")
STATO = os.path.expanduser("~/Library/Application Support/potential/cabru_correzioni_viste.json")

def leggi_correzioni():
    wb = openpyxl.load_workbook(XLS, read_only=True)
    fuori = []
    for ws in wb.worksheets:
        if ws.title == "Istruzioni": continue
        for r in range(3, ws.max_row + 1):
            v = ws.cell(r, 3).value
            if v and str(v).strip():
                fuori.append({"foglio": ws.title, "riga": r,
                              "vecchio": str(ws.cell(r, 2).value or "")[:200],
                              "nuovo": str(v).strip()})
    wb.close()
    return fuori

def impronta(c):
    return hashlib.sha256(f"{c['foglio']}|{c['riga']}|{c['nuovo']}".encode()).hexdigest()[:16]

def main():
    sempre = "--sempre" in sys.argv
    if not os.path.exists(XLS):
        # il file non c'e': e' un guasto vero, non un giorno tranquillo
        print(f"ERRORE: non trovo {XLS}", file=sys.stderr)
        sys.exit(1)

    corr = leggi_correzioni()
    os.makedirs(os.path.dirname(STATO), exist_ok=True)
    try:
        viste = set(json.load(open(STATO, encoding="utf-8"))["viste"])
    except Exception:
        viste = set()

    nuove = [c for c in corr if impronta(c) not in viste]
    oggi = datetime.date.today().strftime("%d.%m.%Y")
    print(f"[{oggi}] correzioni nel foglio: {len(corr)} — di cui nuove: {len(nuove)}")

    if nuove:
        pagine = sorted({c["foglio"] for c in nuove})
        riga = (f"CABRU — {len(nuove)} "
                + ("correzione nuova" if len(nuove) == 1 else "correzioni nuove")
                + " nel foglio dei testi: " + ", ".join(pagine[:4])
                + (f" e altre {len(pagine)-4}" if len(pagine) > 4 else ""))
        dettaglio = "\n".join(f"• [{c['foglio']}] {c['nuovo'][:120]}" for c in nuove[:6])
        if len(nuove) > 6: dettaglio += f"\n… e altre {len(nuove)-6}"
        try:
            import notifica_telegram
            notifica_telegram.notifica(riga + "\n\n" + dettaglio, silenzioso=True)
            print("avviso Telegram mandato")
        except Exception as e:
            # il canale che tace quando si rompe e' il guasto da evitare: l'errore esce comunque
            print(f"Telegram non ha funzionato: {e}", file=sys.stderr)
        json.dump({"viste": sorted(viste | {impronta(c) for c in corr}),
                   "ultimo_giro": oggi},
                  open(STATO, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    elif sempre:
        try:
            import notifica_telegram
            notifica_telegram.notifica(f"CABRU — nessuna correzione nuova nel foglio dei testi ({oggi})", silenzioso=True)
        except Exception as e:
            print(f"Telegram non ha funzionato: {e}", file=sys.stderr)

if __name__ == "__main__":
    main()
