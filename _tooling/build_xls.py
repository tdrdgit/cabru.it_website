#!/usr/bin/env python3
# Estrae tutti i testi italiani dalle pagine del sito e crea un XLS per la correzione.
import sys, os, re, json
HERE = os.path.dirname(os.path.abspath(__file__))
_pylib = os.path.join(HERE, "pylib")
if os.path.isdir(_pylib): sys.path.insert(0, _pylib)  # deps locali se presenti, altrimenti quelle di sistema
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SITE = os.path.dirname(HERE)                # la cartella sito/ (padre di _tooling)
BASE = os.path.dirname(SITE)                # cartella padre del repo: l'XLS finisce qui, FUORI dal sito pubblicato
OUT  = os.path.join(BASE, "CABRU_testi_IT_da_correggere.xlsx")
DESC = json.load(open(os.path.join(HERE, "desc.json"), encoding="utf-8"))["it"]

# ---------- stili ----------
TITLE_FILL = PatternFill("solid", fgColor="0F80A8")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=13)
HEAD_FILL  = PatternFill("solid", fgColor="2B2B30")
HEAD_FONT  = Font(bold=True, color="FFFFFF", size=10)
DIV_FILL   = PatternFill("solid", fgColor="EDEDED")   # divisore grigio chiaro (titoli sezione)
DIV_FONT   = Font(bold=True, color="222222", size=11)
MARK_FILL  = PatternFill("solid", fgColor="D9D9D9")
MARK_FONT  = Font(bold=True, color="333333", size=10)
CORR_FILL  = PatternFill("solid", fgColor="FFFCEC")   # colonna correzione, leggermente evidenziata
BODY_FONT  = Font(size=10, color="222222")
TYPE_FONT  = Font(size=9,  color="777777", italic=True)
WRAP  = Alignment(wrap_text=True, vertical="top")
WRAPC = Alignment(wrap_text=True, vertical="center")
thin  = Side(style="thin", color="DDDDDD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def norm(s):
    return re.sub(r"\s+", " ", (s or "")).strip()

def text_of(el):
    return norm(el.get_text(" ", strip=True))

# ---------- estrazione da pagina generata (article.doc) ----------
SKIP_CLASS = {"cta-block", "brand-hero", "crumb"}
def walk_doc(node, rows):
    for ch in node.children:
        if getattr(ch, "name", None) is None:
            continue
        cls = set(ch.get("class", []))
        if cls & SKIP_CLASS:
            continue
        name = ch.name
        if name in ("script", "style", "svg", "img"):
            continue
        if name in ("h1", "h2", "h3", "h4"):
            t = text_of(ch)
            if not t: continue
            if "faq-q" in cls:
                rows.append(("FAQ — Domanda", t, True)); continue
            typ = {"h1": "Titolo pagina", "h2": "Titolo sezione", "h3": "Sottotitolo", "h4": "Sottotitolo"}[name]
            rows.append((typ, t, name in ("h2", "h3", "h4"))); continue
        if name == "p":
            t = text_of(ch)
            if t: rows.append(("Testo", t, False))
            continue
        if name in ("ul", "ol"):
            for li in ch.find_all("li", recursive=False):
                t = text_of(li)
                if t: rows.append(("Voce elenco", t, False))
            continue
        if name == "div" and "faq-a" in cls:
            t = text_of(ch)
            if t: rows.append(("FAQ — Risposta", t, False))
            continue
        walk_doc(ch, rows)

def extract_doc(path):
    soup = BeautifulSoup(open(path, encoding="utf-8").read(), "html.parser")
    title = norm(soup.title.get_text()) if soup.title else ""
    md = soup.find("meta", attrs={"name": "description"})
    meta = norm(md["content"]) if md and md.get("content") else ""
    art = soup.select_one("article.doc")
    rows = []
    if art:
        walk_doc(art, rows)
    return title, meta, rows

# ---------- estrazione HOME ----------
def walk_home(node, rows):
    for ch in node.children:
        if getattr(ch, "name", None) is None:
            continue
        name = ch.name
        cls = set(ch.get("class", []))
        if name in ("script", "style", "svg", "img", "header", "footer", "nav", "input", "label"):
            continue
        if cls & {"partners-grid", "stats", "lang-switch", "nav-burger", "nav"}:
            continue
        if "eyebrow" in cls:
            t = text_of(ch)
            if t: rows.append(("Occhiello", t, False))
            continue
        if name == "h1":
            rows.append(("Titolo pagina", text_of(ch), False)); continue
        if name in ("h2",):
            rows.append(("Titolo sezione", text_of(ch), True)); continue
        if name in ("h3", "h4"):
            t = text_of(ch)
            if t: rows.append(("Sottotitolo", t, True))
            continue
        if name == "p":
            t = text_of(ch)
            if t: rows.append(("Testo", t, False))
            continue
        if name == "li":
            t = text_of(ch)
            if t: rows.append(("Voce elenco", t, False))
            continue
        if name == "a" and "feat-card" in cls:
            co = ch.select_one(".feat-co"); nm = ch.select_one(".feat-name"); ds = ch.select_one(".feat-desc")
            lab = "Prodotto in evidenza" + (" — " + text_of(co) if co else "")
            val = norm((text_of(nm) if nm else "") + " — " + (text_of(ds) if ds else ""))
            rows.append((lab, val, False)); continue
        if name == "div" and "v" in cls:
            t = text_of(ch)
            if t: rows.append(("Contatto", t, False))
            continue
        walk_home(ch, rows)

def extract_home(path):
    soup = BeautifulSoup(open(path, encoding="utf-8").read(), "html.parser")
    title = norm(soup.title.get_text()) if soup.title else ""
    md = soup.find("meta", attrs={"name": "description"})
    meta = norm(md["content"]) if md and md.get("content") else ""
    rows = []
    walk_home(soup.body, rows)
    return title, meta, rows

# ---------- scrittura foglio ----------
def col_setup(ws):
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95
    ws.column_dimensions["C"].width = 70
    ws.sheet_view.showGridLines = False

def est_height(text, width_chars=95, line_px=15, pad=6):
    if not text: return 16
    lines = 0
    for part in str(text).split("\n"):
        lines += max(1, -(-len(part) // width_chars))
    return min(400, pad + lines * line_px)

def write_sheet(wb, tab, page_title, title_tag, meta, rows, list_mode=False):
    ws = wb.create_sheet(title=tab[:31])
    col_setup(ws)
    r = 1
    # barra titolo pagina
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(r, 1, page_title); c.fill = TITLE_FILL; c.font = TITLE_FONT; c.alignment = WRAPC
    ws.row_dimensions[r].height = 26; r += 1
    # intestazione colonne
    heads = ("Azienda", "Descrizione attuale (IT)", "Correzione (scrivi qui)") if list_mode else \
            ("Tipo", "Testo attuale (IT)", "Correzione (scrivi qui)")
    for j, h in enumerate(heads, 1):
        c = ws.cell(r, j, h); c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = WRAPC; c.border = BORDER
    ws.row_dimensions[r].height = 20; r += 1
    ws.freeze_panes = ws.cell(r, 1)

    def marker(txt):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(r, 1, txt); c.fill = MARK_FILL; c.font = MARK_FONT; c.alignment = WRAP
        ws.row_dimensions[r].height = 18; r += 1

    def row(typ, text, divider=False):
        nonlocal r
        a = ws.cell(r, 1, typ); b = ws.cell(r, 2, text); cc = ws.cell(r, 3, "")
        a.alignment = WRAP; b.alignment = WRAP; cc.alignment = WRAP
        a.border = b.border = cc.border = BORDER
        if divider:
            a.fill = b.fill = DIV_FILL; b.font = DIV_FONT; a.font = TYPE_FONT
        else:
            a.font = TYPE_FONT; b.font = BODY_FONT
        cc.fill = CORR_FILL
        ws.row_dimensions[r].height = est_height(text)
        r += 1

    if not list_mode:
        marker("SEO (testi per Google — titolo e descrizione della pagina)")
        row("SEO — Title", title_tag)
        row("SEO — Meta description", meta)
        marker("CONTENUTO DELLA PAGINA")
    for tup in rows:
        if list_mode:
            row(tup[0], tup[1], False)
        else:
            typ, text, div = tup
            row(typ, text, div)
    return ws

# ---------- costruzione workbook ----------
wb = openpyxl.Workbook()
wb.remove(wb.active)

# foglio istruzioni
ws = wb.create_sheet("Istruzioni")
ws.column_dimensions["A"].width = 110
ws.sheet_view.showGridLines = False
instr = [
 ("CABRU — Testi del sito da correggere", True),
 ("", False),
 ("Come funziona:", True),
 ("• Un foglio (tab) per ogni pagina del sito. Il nome completo della pagina è nella barra blu in alto a ogni foglio.", False),
 ("• Colonna 'Testo attuale (IT)': è quello che c'è oggi online. NON modificarla.", False),
 ("• Colonna 'Correzione (scrivi qui)': scrivi qui la versione corretta SOLO per le righe che vuoi cambiare. Lascia vuoto dove va bene.", False),
 ("• Le righe con sfondo grigio sono i titoli di sezione (fanno da divisori).", False),
 ("• Le prime righe di ogni pagina ('SEO') sono il titolo e la descrizione che compaiono su Google: importanti per il posizionamento.", False),
 ("• Il foglio 'Le aziende che rappresentiamo' elenca nome + descrizione breve di ogni azienda (i loghi non servono).", False),
 ("• Non rinominare i tab: mi servono per rimettere i testi corretti al posto giusto.", False),
 ("", False),
 ("Quando hai finito, rimandami il file: applico le correzioni al sito e ripubblico.", False),
]
rr = 1
for txt, big in instr:
    c = ws.cell(rr, 1, txt); c.alignment = WRAP
    c.font = Font(bold=big, size=13 if big and rr == 1 else (11 if big else 10))
    ws.row_dimensions[rr].height = est_height(txt, 110) if txt else 10
    rr += 1

count = 0

# 1) HOME
t, m, rows = extract_home(os.path.join(SITE, "index.html"))
write_sheet(wb, "Home", "Home — pagina iniziale", t, m, rows); count += 1

# 2) Prodotti hub
t, m, rows = extract_doc(os.path.join(SITE, "prodotti/index.html"))
write_sheet(wb, "Prodotti (hub categorie)", "Prodotti — hub delle categorie", t, m, rows); count += 1

# 3) Categorie
CATS = [
 ("anticorpi", "Anticorpi"),
 ("molecole-biochimiche", "Molecole biochimiche"),
 ("kit-elisa", "Kit ELISA"),
 ("proteine", "Proteine"),
 ("acidi-nucleici", "Acidi nucleici"),
 ("elettroforesi-western-blot", "Elettroforesi WB"),
 ("terreni-microbiologia", "Terreni microbiologia"),
 ("soluzioni-polveri", "Soluzioni e polveri"),
 ("plasmi", "Plasmi"),
 ("prodotti-diagnostici", "Prodotti diagnostici"),
]
for slug, short in CATS:
    p = os.path.join(SITE, "prodotti", slug, "index.html")
    t, m, rows = extract_doc(p)
    ptitle = rows[0][1] if rows and rows[0][0] == "Titolo pagina" else short
    write_sheet(wb, "Prod_" + short, "Prodotti — " + ptitle, t, m, rows); count += 1

# 4) Foglio riepilogo aziende (nome + descrizione)
t, m, hubrows = extract_doc(os.path.join(SITE, "aziende/index.html"))
BRANDS = ["cayman-chemical", "a-a-biotechnology", "affinity-biologicals", "bioatlas",
          "biomedica-diagnostics", "biovendor", "candor-bioscience", "condalab", "dia-pro",
          "enzyme-research-laboratories", "exbio", "finetest", "g-biosciences", "immbiomed",
          "ldn", "magbio", "reliatech", "rovalab", "seqens", "smobio"]
list_rows = [(DESC[s]["name"], DESC[s]["desc"]) for s in BRANDS]
ws = write_sheet(wb, "Le aziende che rappresentiamo",
                 "Le aziende che rappresentiamo — nome + descrizione", "", "",
                 list_rows, list_mode=True)
count += 1

# 5) Pagine azienda (una per brand)
NAMES = {s: DESC[s]["name"] for s in BRANDS}
for s in BRANDS:
    p = os.path.join(SITE, "aziende", s, "index.html")
    t, m, rows = extract_doc(p)
    ptitle = rows[0][1] if rows and rows[0][0] == "Titolo pagina" else NAMES[s]
    tab = "Az_" + NAMES[s]
    write_sheet(wb, tab, "Azienda — " + ptitle, t, m, rows); count += 1

# 6) Guida + applicazione
for path, tab, label in [
    ("guide/eicosanoidi/index.html", "Guida_Eicosanoidi", "Guida"),
    ("applicazioni/tossicologia-forense/index.html", "App_Tossicologia forense", "Applicazione"),
]:
    t, m, rows = extract_doc(os.path.join(SITE, path))
    ptitle = rows[0][1] if rows and rows[0][0] == "Titolo pagina" else tab
    write_sheet(wb, tab, label + " — " + ptitle, t, m, rows); count += 1

# 7) Catalogo (solo intro/testi, non i 118k prodotti)
t, m, rows = extract_doc(os.path.join(SITE, "catalogo/index.html"))
if not rows:
    # il catalogo non usa article.doc: prendo intro h1 + p iniziali
    soup = BeautifulSoup(open(os.path.join(SITE, "catalogo/index.html"), encoding="utf-8").read(), "html.parser")
    t = norm(soup.title.get_text()) if soup.title else ""
    md = soup.find("meta", attrs={"name": "description"})
    m = norm(md["content"]) if md and md.get("content") else ""
    rows = []
    h1 = soup.find("h1")
    if h1: rows.append(("Titolo pagina", text_of(h1), False))
    intro = soup.select_one(".cat-intro, .lead, main p, .wrap > p")
    if intro: rows.append(("Testo", text_of(intro), False))
write_sheet(wb, "Catalogo (intro)", "Catalogo — testi introduttivi", t, m, rows); count += 1

wb.save(OUT)
print("Fogli scritti:", count + 1, "(+ Istruzioni)")
print("Salvato in:", OUT)
# riepilogo tab
print("Tab:", " | ".join(ws.title for ws in wb.worksheets))
