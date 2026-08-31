#!/usr/bin/env python3
"""Le correzioni scritte nella colonna gialla: prima si elencano, poi si chiudono.

  python3 correzioni.py elenca   -> cosa qualcuno ha chiesto di cambiare, e dove
  python3 correzioni.py chiudi   -> DOPO aver messo la modifica nel sito (IT+EN):
                                    porta il testo corretto in colonna B, lo lascia
                                    verde, svuota la colonna C e lo registra.

Il verde dice "questo l'ha voluto un umano": da quel momento un testo verde non si
tocca in autonomia senza segnalarlo, soprattutto se e' una traduzione.
"""
import sys, os, json
HERE = os.path.dirname(os.path.abspath(__file__))
_pylib = os.path.join(HERE, "pylib")
if os.path.isdir(_pylib): sys.path.insert(0, _pylib)
import openpyxl
from openpyxl.styles import Font

XLS   = os.path.join(HERE, "CABRU_testi_IT_da_correggere.xlsx")
VERDI = os.path.join(HERE, "verdi.json")
GREEN = Font(size=10, color="1B5E20", bold=True)

def pendenti(wb):
    for ws in wb.worksheets:
        if ws.title == "Istruzioni": continue
        for r in range(3, ws.max_row + 1):
            nuovo = ws.cell(r, 3).value
            if nuovo and str(nuovo).strip():
                yield ws, r, str(ws.cell(r,1).value or ""), str(ws.cell(r,2).value or ""), str(nuovo).strip()

def elenca():
    wb = openpyxl.load_workbook(XLS)
    n = 0
    for ws, r, tipo, vecchio, nuovo in pendenti(wb):
        n += 1
        print(f"\n[{n}] {ws.title} — riga {r} — {tipo}")
        print(f"    ora dice : {vecchio}")
        print(f"    va messo : {nuovo}")
    print(f"\ncorrezioni in attesa: {n}")
    if n: print("Vanno prima messe nel sito, in italiano E in inglese. Poi: correzioni.py chiudi")
    return n

def chiudi():
    wb = openpyxl.load_workbook(XLS)
    reg = json.load(open(VERDI, encoding="utf-8"))
    testi = reg.setdefault("testi", {})
    n = 0
    for ws, r, tipo, vecchio, nuovo in list(pendenti(wb)):
        ws.cell(r, 2).value = nuovo          # il testo corretto diventa il testo attuale
        ws.cell(r, 2).font  = GREEN          # e resta verde: l'ha deciso un umano
        ws.cell(r, 3).value = None           # la colonna della richiesta torna vuota
        testi.setdefault(ws.title, [])
        if nuovo not in testi[ws.title]: testi[ws.title].append(nuovo)
        if vecchio in testi.get(ws.title, []): testi[ws.title].remove(vecchio)
        print(f"chiusa: {ws.title} riga {r} — {nuovo[:70]}")
        n += 1
    if n:
        wb.save(XLS)
        json.dump(reg, open(VERDI, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"\ncorrezioni chiuse: {n}. Testi verdi in registro: {sum(len(v) for v in testi.values())}")
    return n

def tocco_un_verde(testo):
    """Da chiamare prima di riscrivere un testo: dice se qualcuno l'aveva deciso a mano."""
    reg = json.load(open(VERDI, encoding="utf-8")).get("testi", {})
    for tab, lista in reg.items():
        if testo.strip() in [t.strip() for t in lista]:
            return tab
    return None

if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "elenca"
    if cmd == "elenca": elenca()
    elif cmd == "chiudi": chiudi()
    else: print(__doc__)
